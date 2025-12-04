"""
FastAPI application for AI Data Analyst Pipeline
"""
import asyncio
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import logging

from analyzer import NewsAnalyzer

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="AI Data Analyst API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Storage for tasks
TASKS: Dict[str, Dict] = {}
RESULTS_DIR = Path("/tmp/ai-data-analyst-results")
RESULTS_DIR.mkdir(exist_ok=True)


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "ai-data-analyst"}


@app.post("/api/v1/pipelines/ai-data-analyst/analyze")
async def analyze_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    api_key: str = Form(...),
    model: str = Form(default="gpt-4-turbo-preview")
):
    """
    Upload Excel/CSV file and start analysis
    
    Args:
        file: Excel or CSV file with GNO column
        api_key: OpenAI API key
        model: OpenAI model to use
        
    Returns:
        Task ID for tracking progress
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="File must be Excel (.xlsx, .xls) or CSV (.csv)"
        )
    
    # Generate task ID
    task_id = str(uuid.uuid4())
    
    # Save uploaded file
    upload_path = RESULTS_DIR / f"{task_id}_input{Path(file.filename).suffix}"
    with open(upload_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Initialize task
    TASKS[task_id] = {
        "status": "pending",
        "progress": 0,
        "total": 0,
        "message": "Task queued",
        "created_at": datetime.now().isoformat(),
        "input_file": str(upload_path),
        "output_file": None,
        "error": None
    }
    
    # Start background task
    background_tasks.add_task(
        process_excel_task,
        task_id=task_id,
        input_file=upload_path,
        api_key=api_key,
        model=model
    )
    
    logger.info(f"Created task {task_id} for file {file.filename}")
    
    return {
        "task_id": task_id,
        "message": "Analysis started",
        "status": "pending"
    }


@app.get("/api/v1/pipelines/ai-data-analyst/status/{task_id}")
async def get_task_status(task_id: str):
    """
    Get status of analysis task
    
    Args:
        task_id: Task ID from analyze endpoint
        
    Returns:
        Task status and progress
    """
    if task_id not in TASKS:
        raise HTTPException(status_code=404, detail="Task not found")
    
    return TASKS[task_id]


@app.get("/api/v1/pipelines/ai-data-analyst/download/{task_id}")
async def download_result(task_id: str):
    """
    Download analysis result as Excel file
    
    Args:
        task_id: Task ID from analyze endpoint
        
    Returns:
        Excel file with analysis results
    """
    if task_id not in TASKS:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task = TASKS[task_id]
    
    if task["status"] != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task is not completed. Current status: {task['status']}"
        )
    
    if not task["output_file"] or not Path(task["output_file"]).exists():
        raise HTTPException(status_code=404, detail="Result file not found")
    
    return FileResponse(
        path=task["output_file"],
        filename=f"analysis_result_{task_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


async def process_excel_task(task_id: str, input_file: Path, api_key: str, model: str):
    """
    Background task to process Excel file
    
    Args:
        task_id: Task ID
        input_file: Path to uploaded file
        api_key: OpenAI API key
        model: OpenAI model name
    """
    try:
        # Update status
        TASKS[task_id]["status"] = "processing"
        TASKS[task_id]["message"] = "Reading file..."
        
        # Read Excel/CSV
        if input_file.suffix == '.csv':
            df = pd.read_csv(input_file)
        else:
            df = pd.read_excel(input_file)
        
        logger.info(f"Read file with {len(df)} rows and columns: {list(df.columns)}")
        
        # Find GNO column (case-insensitive)
        gno_column = None
        for col in df.columns:
            if col.lower() in ['gno', 'GNO']:
                gno_column = col
                break
        
        if gno_column is None:
            raise ValueError("Could not find 'Gno' or 'GNO' column in file")
        
        # Extract hyperlinks from Excel cells using openpyxl
        gno_to_url = {}
        if input_file.suffix in ['.xlsx', '.xls']:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(input_file)
                ws = wb.active
                
                # Find GNO column index
                gno_col_idx = None
                for idx, cell in enumerate(ws[1], 1):  # Header row
                    if cell.value and str(cell.value).lower() == 'gno':
                        gno_col_idx = idx
                        break
                
                if gno_col_idx:
                    # Extract hyperlinks from GNO column
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=gno_col_idx, max_col=gno_col_idx), 2):
                        cell = row[0]
                        if cell.value and cell.hyperlink:
                            gno_value = str(cell.value)
                            hyperlink_url = cell.hyperlink.target
                            gno_to_url[gno_value] = hyperlink_url
                            logger.info(f"Found hyperlink for GNO {gno_value}: {hyperlink_url}")
                
                logger.info(f"Extracted {len(gno_to_url)} hyperlinks from Excel")
            except Exception as e:
                logger.warning(f"Could not extract hyperlinks: {str(e)}")
        
        # Get unique GNOs
        unique_gnos = df[gno_column].dropna().unique()
        total_gnos = len(unique_gnos)
        
        TASKS[task_id]["total"] = total_gnos
        TASKS[task_id]["message"] = f"Processing {total_gnos} unique GNOs..."
        
        logger.info(f"Found {total_gnos} unique GNOs")
        
        # Process each GNO
        all_results = []
        
        async with NewsAnalyzer(api_key=api_key, model=model) as analyzer:
            for idx, gno in enumerate(unique_gnos):
                try:
                    TASKS[task_id]["progress"] = idx
                    TASKS[task_id]["message"] = f"Processing GNO {idx + 1}/{total_gnos}: {gno}"
                    
                    logger.info(f"Processing GNO {idx + 1}/{total_gnos}: {gno}")
                    
                    # Use hyperlink if available, otherwise use GNO value directly
                    gno_str = str(gno)
                    gno_url = gno_to_url.get(gno_str, gno_str)
                    
                    if gno_url != gno_str:
                        logger.info(f"Using hyperlink URL: {gno_url}")
                    
                    # Process GNO
                    results = await analyzer.process_gno(gno=gno_str, gno_url=gno_url)
                    all_results.extend(results)
                    
                except Exception as e:
                    logger.error(f"Error processing GNO {gno}: {str(e)}")
                    all_results.append({
                        "gno": str(gno),
                        "error": str(e),
                        "brand": "",
                        "headline": "",
                        "category": "",
                        "sentiment": "",
                        "mention_weight": "",
                        "control": ""
                    })
        
        # Create output DataFrame
        output_df = pd.DataFrame(all_results)
        
        # Reorder columns
        column_order = [
            "gno", "brand", "headline", "category",
            "sentiment", "mention_weight", "control", "error"
        ]
        output_df = output_df[[col for col in column_order if col in output_df.columns]]
        
        # Save to Excel
        output_file = RESULTS_DIR / f"{task_id}_output.xlsx"
        output_df.to_excel(output_file, index=False, engine='openpyxl')
        
        logger.info(f"Saved results to {output_file}")
        
        # Update task
        TASKS[task_id]["status"] = "completed"
        TASKS[task_id]["progress"] = total_gnos
        TASKS[task_id]["message"] = f"Analysis completed. Processed {total_gnos} GNOs, found {len(all_results)} brand mentions."
        TASKS[task_id]["output_file"] = str(output_file)
        TASKS[task_id]["completed_at"] = datetime.now().isoformat()
        
    except Exception as e:
        logger.error(f"Error processing task {task_id}: {str(e)}")
        TASKS[task_id]["status"] = "failed"
        TASKS[task_id]["message"] = f"Error: {str(e)}"
        TASKS[task_id]["error"] = str(e)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8009)
